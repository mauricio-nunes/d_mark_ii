from typing import Iterable, List, Dict, Any
from openpyxl import load_workbook
import warnings
warnings.simplefilter("ignore", UserWarning)  # Suppress openpyxl warnings about data

def read_xlsx_rows(path: str, sheet_name: str | None = None) -> List[Dict[str, Any]]:
    wb = load_workbook(filename=path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []
    for r in rows[1:]:
        row_dict = {}
        for k, v in zip(header, r):
            row_dict[k] = v
        data.append(row_dict)
    return data

def list_sheets(path: str) -> list[str]:
    wb = load_workbook(filename=path, read_only=True)
    return wb.sheetnames
